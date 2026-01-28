"""
Pipeline completo de captacao e qualificacao de leads

Orquestra todos os modulos:
1. Scraping do Google Maps
2. Extracao de redes sociais
3. Enriquecimento (opcional)
4. Scoring
5. Sincronizacao com Airtable
"""
import time
import json
import structlog
from datetime import datetime
from typing import Optional
from pathlib import Path

from config.settings import BUSINESS_TYPES
from src.models import Lead, ScrapingResult
from src.scrapers import GoogleMapsSerpAPI, GoogleMapsScraper
from src.enrichers import SocialMediaExtractor, WebsiteAnalyzer, HunterEnricher
from src.scoring import LeadScorer
from src.integrations import AirtableSync
from src.cache import LeadCache

logger = structlog.get_logger()

CHECKPOINT_FILE = "data/checkpoint.json"


class LeadPipeline:
    """
    Pipeline de captacao e qualificacao de leads

    Fluxo:
    1. Scraping Google Maps -> lista de leads basicos
    2. Analise de websites -> verifica sites ativos
    3. Extracao de redes sociais -> Instagram, LinkedIn
    4. Enriquecimento Hunter.io -> emails, dados extras
    5. Scoring -> pontuacao e classificacao
    6. Sincronizacao Airtable -> persistencia
    """

    def __init__(
        self,
        use_serpapi: bool = True,
        use_hunter: bool = False,
        sync_to_airtable: bool = True,
        use_cache: bool = True,
        use_variations: bool = True,
        max_neighborhoods: int = 5,
    ):
        """
        Inicializa pipeline

        Args:
            use_serpapi: Usar SerpAPI (recomendado) ou scraping direto
            use_hunter: Usar Hunter.io para enriquecimento
            sync_to_airtable: Sincronizar resultados com Airtable
            use_cache: Usar cache para evitar duplicatas entre execucoes
            use_variations: Usar bairros e sinonimos na busca
            max_neighborhoods: Maximo de bairros por categoria
        """
        self.use_serpapi = use_serpapi
        self.use_hunter = use_hunter
        self.sync_to_airtable = sync_to_airtable
        self.use_cache = use_cache
        self.use_variations = use_variations
        self.max_neighborhoods = max_neighborhoods

        # Inicializar componentes
        self._init_components()

    def _init_components(self):
        """Inicializa componentes do pipeline"""
        # Scraper
        if self.use_serpapi:
            try:
                self.scraper = GoogleMapsSerpAPI()
                logger.info("Usando SerpAPI para scraping")
            except ValueError:
                logger.warning("SerpAPI nao configurada, usando scraping direto")
                self.scraper = GoogleMapsScraper()
        else:
            self.scraper = GoogleMapsScraper()
            logger.info("Usando scraping direto")

        # Enrichers
        self.website_analyzer = WebsiteAnalyzer()
        self.social_extractor = SocialMediaExtractor()

        if self.use_hunter:
            try:
                self.hunter = HunterEnricher()
                logger.info("Hunter.io ativado")
            except Exception:
                self.hunter = None
                logger.warning("Hunter.io nao disponivel")
        else:
            self.hunter = None

        # Scorer
        self.scorer = LeadScorer()

        # Airtable
        if self.sync_to_airtable:
            try:
                self.airtable = AirtableSync()
                logger.info("Airtable conectado")
            except Exception as e:
                self.airtable = None
                logger.warning(f"Airtable nao disponivel: {e}")
        else:
            self.airtable = None

        # Cache
        if self.use_cache:
            self.cache = LeadCache()
            stats = self.cache.get_stats()
            logger.info(f"Cache ativo: {stats['total_cached']} leads em cache")
        else:
            self.cache = None

    def _save_checkpoint(self, stage: int, leads: list[Lead], results: dict):
        """Salva checkpoint para poder retomar execucao"""
        checkpoint_path = Path(CHECKPOINT_FILE)
        checkpoint_path.parent.mkdir(parents=True, exist_ok=True)

        leads_data = [lead.model_dump(mode="json") for lead in leads]

        checkpoint = {
            "stage": stage,
            "leads": leads_data,
            "results": results,
            "saved_at": datetime.now().isoformat(),
        }

        with open(checkpoint_path, "w", encoding="utf-8") as f:
            json.dump(checkpoint, f, indent=2, ensure_ascii=False, default=str)

        logger.info(f"Checkpoint salvo: stage {stage}, {len(leads)} leads")

    def _load_checkpoint(self) -> Optional[dict]:
        """Carrega checkpoint anterior se existir"""
        checkpoint_path = Path(CHECKPOINT_FILE)
        if not checkpoint_path.exists():
            return None

        try:
            with open(checkpoint_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            # Reconstituir leads
            leads = [Lead.model_validate(ld) for ld in data.get("leads", [])]
            data["leads"] = leads

            logger.info(
                f"Checkpoint encontrado: stage {data['stage']}, "
                f"{len(leads)} leads, salvo em {data['saved_at']}"
            )
            return data
        except Exception as e:
            logger.warning(f"Erro ao carregar checkpoint: {e}")
            return None

    def _clear_checkpoint(self):
        """Remove checkpoint apos execucao completa"""
        checkpoint_path = Path(CHECKPOINT_FILE)
        if checkpoint_path.exists():
            checkpoint_path.unlink()
            logger.info("Checkpoint removido (execucao completa)")

    def run(
        self,
        categories: Optional[list[str]] = None,
        limit_per_category: int = 20,
        resume: bool = False,
    ) -> dict:
        """
        Executa pipeline completo

        Args:
            categories: Categorias para buscar (default: todas)
            limit_per_category: Max leads por categoria
            resume: Retomar de checkpoint anterior

        Returns:
            Resumo da execucao
        """
        start_time = time.time()
        categories = categories or BUSINESS_TYPES
        resume_stage = 0
        leads = []

        # Tentar retomar de checkpoint
        if resume:
            checkpoint = self._load_checkpoint()
            if checkpoint:
                resume_stage = checkpoint["stage"]
                leads = checkpoint["leads"]
                results = checkpoint.get("results", {
                    "started_at": datetime.now().isoformat(),
                    "categories": categories,
                    "stages": {},
                })
                logger.info(
                    f"Retomando pipeline do stage {resume_stage + 1} "
                    f"com {len(leads)} leads"
                )
            else:
                logger.info("Nenhum checkpoint encontrado, iniciando do zero")
                resume = False

        if not resume:
            results = {
                "started_at": datetime.now().isoformat(),
                "categories": categories,
                "stages": {},
            }

        logger.info(
            f"Iniciando pipeline: {len(categories)} categorias, "
            f"limit={limit_per_category}"
        )

        # Testar conexao Airtable ANTES de gastar API calls
        if self.airtable:
            logger.info("Testando conexao com Airtable...")
            if not self.airtable.test_connection():
                logger.error(
                    "ABORTANDO: Airtable sem permissao. "
                    "Corrija as permissoes do token antes de executar. "
                    "Use --no-airtable para rodar sem sincronizar."
                )
                results["error"] = "Airtable permission denied (403)"
                return results

        # Stage 1: Scraping
        if resume_stage < 1:
            logger.info("=== Stage 1: Scraping Google Maps ===")
            leads = self.scraper.search_all_categories(
                categories, limit_per_category,
                use_variations=self.use_variations,
                max_neighborhoods=self.max_neighborhoods,
            )
            total_scraped = len(leads)
            results["stages"]["scraping"] = {
                "leads_found": total_scraped,
            }
            logger.info(f"Scraping: {total_scraped} leads encontrados")

            if not leads:
                logger.warning("Nenhum lead encontrado, encerrando")
                return results

            # Filtrar duplicatas usando cache
            if self.cache:
                leads = self.cache.filter_new(leads)
                results["stages"]["scraping"]["new_leads"] = len(leads)
                results["stages"]["scraping"]["cached_skipped"] = total_scraped - len(leads)

                if not leads:
                    logger.info("Todos os leads ja estao em cache")
                    results["total_leads"] = 0
                    return results

            self._save_checkpoint(1, leads, results)

        # Stage 2: Analise de websites
        if resume_stage < 2:
            logger.info("=== Stage 2: Analise de Websites ===")
            leads = self.website_analyzer.analyze_leads(leads)
            sites_ativos = sum(1 for l in leads if l.site_ativo)
            results["stages"]["website_analysis"] = {
                "sites_ativos": sites_ativos,
                "sites_https": sum(1 for l in leads if l.site_https),
            }
            logger.info(f"Websites: {sites_ativos} sites ativos")
            self._save_checkpoint(2, leads, results)

        # Stage 3: Extracao de redes sociais, emails e telefones
        if resume_stage < 3:
            logger.info("=== Stage 3: Extracao de Redes Sociais, Emails e Telefones ===")
            leads = self.social_extractor.enrich_leads(leads)

            instagram_count = sum(1 for l in leads if l.social.instagram)
            linkedin_count = sum(1 for l in leads if l.social.linkedin)
            email_count = sum(1 for l in leads if l.email)
            telefone_count = sum(1 for l in leads if l.telefone)

            results["stages"]["social_extraction"] = {
                "instagram": instagram_count,
                "linkedin": linkedin_count,
                "emails": email_count,
                "telefones": telefone_count,
            }
            logger.info(
                f"Redes sociais: {instagram_count} Instagram, "
                f"{linkedin_count} LinkedIn"
            )
            logger.info(
                f"Contato: {email_count} emails, {telefone_count} telefones"
            )
            self._save_checkpoint(3, leads, results)

        # Stage 4: Enriquecimento Hunter.io (opcional)
        if resume_stage < 4:
            if self.hunter:
                logger.info("=== Stage 4: Enriquecimento Hunter.io ===")
                leads = self.hunter.enrich_leads(leads)
                emails_found = sum(1 for l in leads if l.email)
                results["stages"]["hunter_enrichment"] = {
                    "emails_found": emails_found,
                }
                logger.info(f"Hunter.io: {emails_found} emails encontrados")
            self._save_checkpoint(4, leads, results)

        # Stage 5: Scoring
        if resume_stage < 5:
            logger.info("=== Stage 5: Scoring ===")
            leads = self.scorer.score_leads(leads)
            summary = self.scorer.get_summary(leads)
            results["stages"]["scoring"] = summary
            logger.info(
                f"Scoring: score medio={summary['score_medio']:.1f}, "
                f"hot={summary['hot_leads']}, warm={summary['warm_leads']}"
            )
            self._save_checkpoint(5, leads, results)

        # Stage 6: Sincronizacao Airtable
        if self.airtable:
            logger.info("=== Stage 6: Sincronizacao Airtable ===")
            sync_result = self.airtable.sync_leads(leads)
            results["stages"]["airtable_sync"] = sync_result
            logger.info(
                f"Airtable: {sync_result['created']} criados, "
                f"{sync_result['updated']} atualizados"
            )

        # Salvar leads no cache - SOMENTE leads confirmados no Airtable
        # Se Airtable esta ativo, so cacheia leads que foram sincronizados com sucesso
        # Se Airtable esta desativado, cacheia todos (comportamento anterior)
        if self.cache:
            if self.airtable:
                synced_leads = [l for l in leads if l.synced_to_airtable]
                if synced_leads:
                    self.cache.add_many(synced_leads)
                    logger.info(
                        f"Cache atualizado: {len(synced_leads)} leads sincronizados "
                        f"(total cache: {self.cache.get_stats()['total_cached']})"
                    )
                failed_leads = [l for l in leads if not l.synced_to_airtable]
                if failed_leads:
                    logger.warning(
                        f"{len(failed_leads)} leads NAO cacheados "
                        f"(falha no Airtable - serao reprocessados na proxima execucao)"
                    )
            else:
                self.cache.add_many(leads)
                logger.info(f"Cache atualizado: {self.cache.get_stats()['total_cached']} leads")

        # Stage 7: Exportar Excel
        logger.info("=== Stage 7: Exportacao Excel ===")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = f"data/leads_{timestamp}.xlsx"
        Path("data").mkdir(parents=True, exist_ok=True)
        self.export_to_excel(leads, excel_path)
        results["excel_file"] = excel_path

        # Finalizar
        duration = time.time() - start_time
        results["duration_seconds"] = duration
        results["finished_at"] = datetime.now().isoformat()
        results["total_leads"] = len(leads)

        # Limpar checkpoint apos execucao completa com sucesso
        self._clear_checkpoint()

        logger.info(f"Pipeline concluido em {duration:.1f}s")
        logger.info(f"Total: {len(leads)} leads processados")

        return results

    def run_single_category(
        self, category: str, limit: int = 20
    ) -> list[Lead]:
        """
        Processa uma unica categoria

        Util para testes e debugging
        """
        return self.run(categories=[category], limit_per_category=limit)

    def export_to_csv(self, leads: list[Lead], filepath: str):
        """Exporta leads para CSV"""
        import csv

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Header
            writer.writerow([
                "Nome", "Categoria", "Telefone", "Email",
                "Endereco", "Site", "Instagram", "LinkedIn",
                "Rating", "Reviews", "Score", "Classificacao"
            ])

            # Data
            for lead in leads:
                writer.writerow([
                    lead.nome,
                    lead.categoria,
                    lead.telefone or "",
                    lead.email or "",
                    lead.endereco or "",
                    lead.site or "",
                    lead.social.instagram or "",
                    lead.social.linkedin or "",
                    lead.google_maps.rating or "",
                    lead.google_maps.num_reviews or "",
                    lead.score,
                    lead.classificacao,
                ])

        logger.info(f"CSV exportado para {filepath}")

    def export_to_excel(self, leads: list[Lead], filepath: str):
        """
        Exporta leads para Excel (.xlsx) com formatacao profissional

        Gera planilha com:
        - Abas separadas: Todos os Leads, Hot, Warm, Cold
        - Cores por classificacao
        - Filtros automaticos
        - Largura de colunas ajustada
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error(
                "openpyxl nao instalado. Instale com: pip install openpyxl"
            )
            # Fallback para CSV
            csv_path = filepath.replace(".xlsx", ".csv")
            logger.info(f"Exportando como CSV: {csv_path}")
            self.export_to_csv(leads, csv_path)
            return csv_path

        wb = Workbook()

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        hot_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        hot_font = Font(bold=True, color="FFFFFF")
        warm_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        warm_font = Font(bold=True, color="FFFFFF")
        cold_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        cold_font = Font(bold=True, color="FFFFFF")
        low_fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
        low_font = Font(color="FFFFFF")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        columns = [
            ("Nome", 35),
            ("Categoria", 22),
            ("Classificacao", 15),
            ("Score", 10),
            ("Telefone", 18),
            ("Email", 30),
            ("Endereco", 40),
            ("Cidade", 18),
            ("Site", 35),
            ("Instagram", 30),
            ("LinkedIn", 30),
            ("Rating", 10),
            ("Num Reviews", 13),
            ("Status", 15),
            ("Data Captura", 18),
        ]

        def _lead_to_row(lead: Lead) -> list:
            return [
                lead.nome,
                lead.categoria,
                lead.classificacao,
                lead.score,
                lead.telefone or "",
                lead.email or "",
                lead.endereco or "",
                lead.cidade,
                lead.site or "",
                lead.social.instagram or "",
                lead.social.linkedin or "",
                lead.google_maps.rating or "",
                lead.google_maps.num_reviews or "",
                lead.status,
                lead.data_captura.strftime("%Y-%m-%d %H:%M") if lead.data_captura else "",
            ]

        def _write_sheet(ws, sheet_leads: list[Lead]):
            # Header
            for col_idx, (col_name, col_width) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = col_width

            # Dados
            for row_idx, lead in enumerate(sheet_leads, 2):
                row_data = _lead_to_row(lead)
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

                # Cor da classificacao (coluna 3)
                class_cell = ws.cell(row=row_idx, column=3)
                score_cell = ws.cell(row=row_idx, column=4)
                if lead.classificacao == "hot":
                    class_cell.fill = hot_fill
                    class_cell.font = hot_font
                    score_cell.fill = hot_fill
                    score_cell.font = hot_font
                elif lead.classificacao == "warm":
                    class_cell.fill = warm_fill
                    class_cell.font = warm_font
                    score_cell.fill = warm_fill
                    score_cell.font = warm_font
                elif lead.classificacao == "cold":
                    class_cell.fill = cold_fill
                    class_cell.font = cold_font
                elif lead.classificacao == "low":
                    class_cell.fill = low_fill
                    class_cell.font = low_font

            # Filtros automaticos
            if sheet_leads:
                last_col = get_column_letter(len(columns))
                ws.auto_filter.ref = f"A1:{last_col}{len(sheet_leads) + 1}"

            # Congelar primeira linha
            ws.freeze_panes = "A2"

        # Aba 1: Todos os leads (ordenados por score decrescente)
        ws_all = wb.active
        ws_all.title = "Todos os Leads"
        sorted_leads = sorted(leads, key=lambda l: l.score, reverse=True)
        _write_sheet(ws_all, sorted_leads)

        # Aba 2: Hot leads
        hot_leads = [l for l in sorted_leads if l.classificacao == "hot"]
        if hot_leads:
            ws_hot = wb.create_sheet("Hot Leads")
            _write_sheet(ws_hot, hot_leads)

        # Aba 3: Warm leads
        warm_leads = [l for l in sorted_leads if l.classificacao == "warm"]
        if warm_leads:
            ws_warm = wb.create_sheet("Warm Leads")
            _write_sheet(ws_warm, warm_leads)

        # Aba 4: Resumo
        ws_summary = wb.create_sheet("Resumo")
        summary_data = [
            ("Metrica", "Valor"),
            ("Total de Leads", len(leads)),
            ("Leads Hot", len(hot_leads)),
            ("Leads Warm", len(warm_leads)),
            ("Leads Cold", len([l for l in leads if l.classificacao == "cold"])),
            ("Leads Low", len([l for l in leads if l.classificacao == "low"])),
            ("", ""),
            ("Score Medio", round(sum(l.score for l in leads) / len(leads), 1) if leads else 0),
            ("Com Telefone", sum(1 for l in leads if l.telefone)),
            ("Com Email", sum(1 for l in leads if l.email)),
            ("Com Site", sum(1 for l in leads if l.site)),
            ("Com Instagram", sum(1 for l in leads if l.social.instagram)),
            ("Com LinkedIn", sum(1 for l in leads if l.social.linkedin)),
        ]

        for row_idx, (metric, value) in enumerate(summary_data, 1):
            cell_a = ws_summary.cell(row=row_idx, column=1, value=metric)
            cell_b = ws_summary.cell(row=row_idx, column=2, value=value)
            cell_a.border = thin_border
            cell_b.border = thin_border
            if row_idx == 1:
                cell_a.font = header_font
                cell_a.fill = header_fill
                cell_b.font = header_font
                cell_b.fill = header_fill

        ws_summary.column_dimensions["A"].width = 20
        ws_summary.column_dimensions["B"].width = 15

        # Salvar
        wb.save(filepath)
        logger.info(
            f"Excel exportado: {filepath} "
            f"({len(leads)} leads, {len(hot_leads)} hot, {len(warm_leads)} warm)"
        )
        return filepath

        logger.info(f"Exportado para {filepath}")


def run_daily_pipeline():
    """
    Funcao para execucao diaria do pipeline

    Pode ser chamada por cron ou N8N
    """
    pipeline = LeadPipeline(
        use_serpapi=True,
        use_hunter=False,  # Economizar creditos
        sync_to_airtable=True,
    )

    results = pipeline.run(limit_per_category=20)

    return results


if __name__ == "__main__":
    # Execucao de teste
    import json

    pipeline = LeadPipeline(
        use_serpapi=True,
        use_hunter=False,
        sync_to_airtable=False,  # Desativar para teste
    )

    # Testar com uma categoria
    results = pipeline.run(
        categories=["clinica medica"],
        limit_per_category=5,
    )

    print(json.dumps(results, indent=2, default=str))
